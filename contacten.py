"""
contacten.py — Blueprint voor de Contacten-module.

Bevat: /contacten (GET/POST) — contactpersonen bij bedrijven bekijken,
toevoegen, verwijderen, doorzoeken op naam/bedrijf/e-mail/accountmanager.

Registratie in app.py met: app.register_blueprint(contacten_bp)
"""
import uuid
import datetime
import io
from flask import Blueprint, request, session, redirect, url_for, render_template_string, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd

from core import (
    laad_contactpersonen, bewaar_contactpersonen, laad_accountmanagers, bewaar_accountmanagers,
    is_huidige_gebruiker_admin, ENF_BEDRIJVEN, render_simple_page, vereist_afdeling_of_403,
    bewaar_bedrijven, PAPIERFABRIEKEN, laad_status, bewaar_status, laad_users, geocode_adres,
)

contacten_bp = Blueprint("contacten", __name__)

@contacten_bp.route("/contacten", methods=["GET", "POST"])
def contacten():
    _guard = vereist_afdeling_of_403("contacten")
    if _guard: return _guard
    if request.method == "POST":
        actie = request.form.get("actie", "")
        personen = laad_contactpersonen()

        if actie == "toevoegen":
            nieuw = {
                "id": str(uuid.uuid4()),
                "naam": request.form.get("naam", "").strip(),
                "bedrijf": request.form.get("bedrijf", "").strip(),
                "rol": request.form.get("rol", "").strip(),
                "email": request.form.get("email", "").strip(),
                "telefoon": request.form.get("telefoon", "").strip(),
                "laatst": request.form.get("laatst", "") or datetime.date.today().isoformat(),
                "gebruiker": session.get("gebruikersnaam", ""),
                "aangemaakt": datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
            }
            if nieuw["naam"] and nieuw["bedrijf"]:
                personen.append(nieuw)
                bewaar_contactpersonen(personen)

        elif actie == "verwijderen":
            persoon_id = request.form.get("persoon_id", "")
            doel = next((p for p in personen if p["id"] == persoon_id), None)
            if doel and (doel.get("gebruiker") == session.get("gebruikersnaam","") or is_huidige_gebruiker_admin()):
                personen = [p for p in personen if p["id"] != persoon_id]
                bewaar_contactpersonen(personen)

        return redirect(url_for("contacten.contacten", **request.args))

    zoekterm = request.args.get("zoekterm", "").strip().lower()
    gekozen_am = request.args.get("accountmanager", "")

    accountmanagers_alle = laad_accountmanagers()
    _bedrijven_land_lookup = {b["naam"]: b.get("land","") for b in ENF_BEDRIJVEN}

    personen = laad_contactpersonen()
    bekende_namen_bedrijven = {(p["naam"].strip().lower(), p["bedrijf"].strip().lower()) for p in personen}

    contacten_lijst = []
    for p in personen:
        contacten_lijst.append({
            "id": p["id"], "naam": p["naam"], "bedrijf": p["bedrijf"], "rol": p.get("rol",""),
            "email": p.get("email",""), "telefoon": p.get("telefoon",""),
            "accountmanager": accountmanagers_alle.get(p["bedrijf"], ""), "laatst": p.get("laatst",""),
            "eigen": p.get("gebruiker") == session.get("gebruikersnaam",""),
        })

    # Bedrijven met een los ingevuld 'contactpersoon'-veld die nog geen formeel record hebben: ook tonen (niet verzinnen, wel niet verliezen)
    for b in ENF_BEDRIJVEN:
        naam_veld = (b.get("contactpersoon","") or "").strip()
        if naam_veld and (naam_veld.lower(), b["naam"].strip().lower()) not in bekende_namen_bedrijven:
            contacten_lijst.append({
                "id": "", "naam": naam_veld, "bedrijf": b["naam"], "rol": "", "email": "", "telefoon": b.get("telefoon",""),
                "accountmanager": accountmanagers_alle.get(b["naam"], ""), "laatst": "", "eigen": False,
            })

    if zoekterm:
        contacten_lijst = [c for c in contacten_lijst if zoekterm in c["naam"].lower() or zoekterm in c["bedrijf"].lower() or zoekterm in c["email"].lower()]
    if gekozen_am:
        contacten_lijst = [c for c in contacten_lijst if c["accountmanager"] == gekozen_am]
    contacten_lijst.sort(key=lambda c: c["naam"])

    alle_accountmanagers = sorted({v for v in accountmanagers_alle.values() if v})

    inhoud = """
<style>
.data-thead, .data-row { display: flex; align-items: center; padding: 0 var(--space-4); }
.data-thead { padding-top: 10px; padding-bottom: 10px; background: var(--gray-50); border-bottom: 1px solid var(--gray-200); font-size: 10px; letter-spacing: 0.08em; text-transform: uppercase; color: #7d8792; }
.data-thead span[data-sort] { cursor: pointer; user-select: none; }
.data-thead span[data-sort]:hover { color: var(--brand-600); }
.data-row { padding-top: 11px; padding-bottom: 11px; border-bottom: 1px solid var(--gray-100); font-size: 12.5px; text-decoration: none; color: inherit; }
.data-row:hover { background: #f9fbfc; }
.data-row .zacht { color: #4b5563; font-size: 12px; }
</style>

<div class="page-title">Contacten</div>
<p style="color:var(--gray-400);margin-top:0;margin-bottom:16px;font-size:0.85rem;">Contactpersonen bij bedrijven, met rol en laatste contactmoment</p>

<form method="GET" style="display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;align-items:center;">
    <input type="text" name="zoekterm" value="{{ zoekterm }}" placeholder="Naam, bedrijf of e-mail" style="flex:1;max-width:280px;padding:7px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:12.5px;font-family:inherit;">
    <select name="accountmanager" onchange="this.form.submit()" style="padding:7px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:12.5px;">
        <option value="">Alle accountmanagers</option>
        {% for a in alle_accountmanagers %}<option value="{{ a }}" {% if gekozen_am == a %}selected{% endif %}>{{ a }}</option>{% endfor %}
    </select>
    <button type="submit" class="btn-nav btn-nav-primary" style="border:none;cursor:pointer;">Zoeken</button>
    {% if zoekterm or gekozen_am %}<a href="/contacten" style="font-size:12px;color:var(--gray-400);text-decoration:none;">Wis filters</a>{% endif %}
</form>

<div style="margin-bottom:20px;">
    <a href="/contacten/nieuw" style="display:inline-block;padding:9px 18px;background:var(--brand-600);color:#fff;text-decoration:none;border-radius:6px;font-size:13px;font-weight:700;">+ Contact toevoegen</a>
</div>

{% if contacten_lijst %}
<div class="data-thead" style="border-radius:var(--radius-md) var(--radius-md) 0 0;">
    <span style="flex:1.2;" data-sort="naam">Contactpersoon</span>
    <span style="flex:1.4;" data-sort="bedrijf">Bedrijf</span>
    <span style="flex:1;" data-sort="rol">Rol</span>
    <span style="flex:1.2;" data-sort="email">E-mail</span>
    <span style="width:130px;" data-sort="telefoon">Telefoon</span>
    <span style="width:110px;" data-sort="accountmanager">Accountmgr.</span>
    <span style="width:90px;text-align:right;" data-sort="laatst">Contact</span>
    <span style="width:26px;"></span>
</div>
<div id="contactenLijst" style="border:1px solid var(--gray-200);border-top:none;border-radius:0 0 var(--radius-md) var(--radius-md);overflow:hidden;">
    {% for c in contacten_lijst %}
    <div class="data-row"
       data-naam="{{ c.naam|e }}" data-bedrijf="{{ c.bedrijf|e }}" data-rol="{{ c.rol|default('',true)|e }}"
       data-email="{{ c.email|default('',true)|e }}" data-telefoon="{{ c.telefoon|default('',true)|e }}"
       data-accountmanager="{{ c.accountmanager|default('',true)|e }}" data-laatst="{{ c.laatst|default('',true)|e }}">
        <span style="flex:1.2;"><a href="/bedrijf/{{ c.bedrijf|urlencode }}" style="font-weight:600;color:var(--gray-800);text-decoration:none;">{{ c.naam }}</a></span>
        <span style="flex:1.4;" class="zacht">{{ c.bedrijf }}</span>
        <span style="flex:1;" class="zacht">{{ c.rol|default('—',true) }}</span>
        <span style="flex:1.2;" class="zacht" style="color:var(--brand-600);">{{ c.email|default('—',true) }}</span>
        <span style="width:130px;" class="num">{{ c.telefoon|default('—',true) }}</span>
        <span style="width:110px;" class="zacht">{{ c.accountmanager|default('—',true) }}</span>
        <span style="width:90px;text-align:right;font-size:11.5px;color:var(--gray-400);">{{ c.laatst|default('—',true) }}</span>
        <span style="width:26px;">
            {% if c.id and c.eigen %}
            <form method="POST" onsubmit="return confirm('Contactpersoon verwijderen?');" style="margin:0;">
                <input type="hidden" name="actie" value="verwijderen"><input type="hidden" name="persoon_id" value="{{ c.id }}">
                <button type="submit" style="background:none;border:none;color:var(--gray-300);cursor:pointer;">✕</button>
            </form>
            {% endif %}
        </span>
    </div>
    {% endfor %}
</div>
<div style="display:flex;justify-content:space-between;padding:10px 4px;font-size:0.8rem;color:var(--gray-400);">
    <span>{{ contacten_lijst|length }} contactpersonen</span>
    <a href="/export-csv" style="color:var(--brand-600);text-decoration:none;font-weight:600;">Export naar CSV</a>
</div>
<script>
(function () {
    var lijst = document.getElementById("contactenLijst");
    if (!lijst) return;
    var koppen = document.querySelectorAll(".data-thead [data-sort]");
    var richting = "desc", sleutel = null;
    koppen.forEach(function (kop) {
        kop.addEventListener("click", function () {
            var k = kop.dataset.sort;
            richting = (sleutel === k && richting === "desc") ? "asc" : "desc";
            sleutel = k;
            koppen.forEach(function (x) { x.textContent = x.textContent.replace(/ [\u2191\u2193]$/, ""); });
            kop.textContent += richting === "desc" ? " \u2193" : " \u2191";
            var rijen = Array.prototype.slice.call(lijst.querySelectorAll(".data-row"));
            rijen.sort(function (a, b) {
                var va = a.dataset[k] || "", vb = b.dataset[k] || "";
                var r = va.localeCompare(vb, "nl");
                return richting === "asc" ? r : -r;
            });
            rijen.forEach(function (r) { lijst.appendChild(r); });
        });
    });
})();
</script>
{% else %}
<div class="lege-staat">Geen contactpersonen gevonden.</div>
{% endif %}
    """
    pagina = render_simple_page("Contacten", "contacten", inhoud)
    return render_template_string(pagina, contacten_lijst=contacten_lijst, zoekterm=zoekterm, gekozen_am=gekozen_am,
                                    alle_accountmanagers=alle_accountmanagers, bedrijfnamen_lijst=sorted(_bedrijven_land_lookup.keys())[:500])

@contacten_bp.route("/contacten/import-sjabloon")
def contacten_import_sjabloon():
    """Downloadbaar .xlsx-sjabloon voor het bulk-importeren van bedrijven —
    bedoeld voor een migratie vanuit een ander systeem (bv. Zoho): daar
    exporteren, hier de kolommen invullen, en uploaden bij 'Bedrijven
    importeren'."""
    _guard = vereist_afdeling_of_403("contacten")
    if _guard: return _guard

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bedrijven"

    kolommen = ["Bedrijfsnaam", "Status (Klant/Leverancier)", "Land", "Stad/Regio", "Materialen",
                "Contactpersoon naam", "Contactpersoon functie", "Contactpersoon e-mail",
                "Contactpersoon telefoon", "Accountmanager (gebruikersnaam)"]
    for kolom_idx, titel in enumerate(kolommen, start=1):
        cel = ws.cell(row=1, column=kolom_idx, value=titel)
        cel.font = Font(name="Arial", bold=True, color="FFFFFF")
        cel.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ws.column_dimensions[cel.column_letter].width = 24

    voorbeeldrij = ["Voorbeeld Papier BV", "Leverancier", "Netherlands", "Rotterdam", "Karton, OCC",
                     "Jan Jansen", "Inkoper", "jan@voorbeeld.nl", "+31 6 12345678", ""]
    for kolom_idx, waarde in enumerate(voorbeeldrij, start=1):
        cel = ws.cell(row=2, column=kolom_idx, value=waarde)
        cel.font = Font(name="Arial", italic=True, color="9CA3AF")

    ws.cell(row=4, column=1, value="Toelichting:").font = Font(name="Arial", bold=True)
    toelichting = [
        "- Alleen 'Bedrijfsnaam' is verplicht — de rest mag leeg blijven.",
        "- Status: precies 'Klant' of 'Leverancier' (leeg = geen van beide, alleen vastgelegd als bedrijf).",
        "- Bedrijven die al bestaan (op naam) worden bij import overgeslagen, niet overschreven.",
        "- Contactpersoon-kolommen zijn optioneel; vul je een naam in, dan wordt die contactpersoon ook aangemaakt.",
        "- Accountmanager moet een bestaande gebruikersnaam in het systeem zijn, anders wordt die kolom genegeerd.",
        "- Verwijder de voorbeeldrij (rij 2) voordat je uploadt, of laat 'm staan — die wordt bij import automatisch herkend en overgeslagen.",
    ]
    for i, regel in enumerate(toelichting, start=5):
        ws.cell(row=i, column=1, value=regel).font = Font(name="Arial", color="6B7280", italic=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="bedrijven_import_sjabloon.xlsx",
                       mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _valideer_import_rij(rij, bestaande_namen_laag, geziene_namen_in_bestand):
    """Eén rij uit het geüploade bestand controleren. Geeft (status, melding, schone_data)
    terug — status is 'ok', 'bestaat_al', 'dubbel_in_bestand', 'voorbeeldrij' of 'geen_naam'."""
    naam = str(rij.get("Bedrijfsnaam", "") or "").strip()
    if not naam or naam.lower() == "nan":
        return "geen_naam", "Geen bedrijfsnaam ingevuld", None
    if naam == "Voorbeeld Papier BV":
        return "voorbeeldrij", "Voorbeeldrij uit het sjabloon", None
    if naam.lower() in bestaande_namen_laag:
        return "bestaat_al", f"'{naam}' bestaat al — overgeslagen (niet overschreven)", None
    if naam.lower() in geziene_namen_in_bestand:
        return "dubbel_in_bestand", f"'{naam}' komt dubbel voor in het bestand — alleen de eerste rij wordt gebruikt", None

    def _schoon(veld):
        w = rij.get(veld, "")
        w = "" if w is None else str(w).strip()
        return "" if w.lower() == "nan" else w

    status_ruw = _schoon("Status (Klant/Leverancier)").lower()
    status = "klant" if status_ruw == "klant" else ("leverancier" if status_ruw == "leverancier" else "")

    accountmanager = _schoon("Accountmanager (gebruikersnaam)")
    alle_gebruikers = laad_users()
    if accountmanager and accountmanager not in alle_gebruikers:
        accountmanager = ""  # onbekende gebruikersnaam wordt genegeerd, geen harde fout

    schone_data = {
        "naam": naam, "status": status, "land": _schoon("Land"), "regio": _schoon("Stad/Regio"),
        "materialen": _schoon("Materialen"), "accountmanager": accountmanager,
        "contact_naam": _schoon("Contactpersoon naam"), "contact_functie": _schoon("Contactpersoon functie"),
        "contact_email": _schoon("Contactpersoon e-mail"), "contact_telefoon": _schoon("Contactpersoon telefoon"),
    }
    return "ok", "Wordt geïmporteerd", schone_data

@contacten_bp.route("/contacten/importeren", methods=["GET", "POST"])
def contacten_importeren():
    """Bedrijven in bulk importeren via een geüpload .xlsx-bestand — voor een
    migratie vanuit een ander systeem (bv. Zoho). Toont eerst een preview met
    per rij wat er gebeurt, voordat er daadwerkelijk iets wordt opgeslagen."""
    _guard = vereist_afdeling_of_403("contacten")
    if _guard: return _guard

    fout = None
    preview_rijen = None
    aantal_ok = 0

    if request.method == "POST":
        bestand = request.files.get("bestand")
        if not bestand or not bestand.filename:
            fout = "Kies eerst een .xlsx-bestand."
        elif not bestand.filename.lower().endswith((".xlsx", ".xls")):
            fout = "Alleen .xlsx- of .xls-bestanden worden ondersteund."
        else:
            try:
                df = pd.read_excel(bestand, dtype=str)
            except Exception:
                fout = "Kon het bestand niet lezen — is het een geldig Excel-bestand?"
            else:
                bestaande_namen_laag = {b["naam"].strip().lower() for b in ENF_BEDRIJVEN}
                geziene_namen_in_bestand = set()
                preview_rijen = []
                for _, rij in df.iterrows():
                    status, melding, schone_data = _valideer_import_rij(rij, bestaande_namen_laag, geziene_namen_in_bestand)
                    if status == "ok":
                        aantal_ok += 1
                        geziene_namen_in_bestand.add(schone_data["naam"].lower())
                    preview_rijen.append({"status": status, "melding": melding, "data": schone_data,
                                            "naam_ruw": str(rij.get("Bedrijfsnaam","") or "")})
                if not preview_rijen:
                    fout = "Het bestand bevat geen rijen."

    inhoud = """
<div class="page-title">Bedrijven importeren</div>
<p style="color:var(--gray-400);margin-top:0;margin-bottom:20px;font-size:0.85rem;">Bulk-importeren vanuit een Excel-bestand — handig bij een migratie vanuit een ander systeem (bv. Zoho): daar exporteren, hier de kolommen invullen en uploaden.</p>

{% if fout %}<div style="background:#fef2f2;color:#dc2626;padding:10px 14px;border-radius:8px;margin-bottom:16px;font-size:12.5px;">{{ fout }}</div>{% endif %}

{% if not preview_rijen %}
<div style="max-width:520px;background:var(--gray-50);border-radius:10px;padding:18px;margin-bottom:20px;">
    <a href="/contacten/import-sjabloon" style="font-size:13px;font-weight:700;color:var(--brand-600);text-decoration:none;">⬇ Download het Excel-sjabloon</a>
    <p style="font-size:12px;color:var(--gray-500);margin:8px 0 0 0;">Vul de kolommen in (alleen bedrijfsnaam is verplicht) en upload het bestand hieronder.</p>
</div>
<form method="POST" enctype="multipart/form-data" style="max-width:520px;">
    <input type="file" name="bestand" accept=".xlsx,.xls" required style="margin-bottom:12px;">
    <br>
    <button type="submit" style="padding:9px 20px;background:var(--brand-600);color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer;">Bestand controleren</button>
</form>
{% else %}
<div style="margin-bottom:16px;font-size:13px;color:var(--gray-700);">
    <b>{{ aantal_ok }}</b> van de {{ preview_rijen|length }} rijen worden geïmporteerd.
    {% if aantal_ok < preview_rijen|length %}De rest wordt overgeslagen (zie reden per rij hieronder).{% endif %}
</div>
<div style="border:none;border-top:1px solid var(--gray-200);border-bottom:1px solid var(--gray-200);margin-bottom:20px;max-width:900px;">
    {% for r in preview_rijen %}
    <div style="display:flex;align-items:center;padding:9px 12px;border-bottom:1px solid var(--gray-100);font-size:12.5px;">
        <span style="width:22px;">{% if r.status == "ok" %}✅{% else %}⚠️{% endif %}</span>
        <span style="flex:1;font-weight:600;color:var(--gray-800);">{{ r.naam_ruw or "(leeg)" }}</span>
        <span style="flex:1.5;color:{% if r.status == 'ok' %}var(--gray-400){% else %}#dc2626{% endif %};">{{ r.melding }}</span>
    </div>
    {% endfor %}
</div>
{% if aantal_ok > 0 %}
<form method="POST" action="/contacten/importeren/bevestigen">
    <input type="hidden" name="import_data" value='{{ te_importeren|tojson }}'>
    <button type="submit" style="padding:9px 20px;background:var(--brand-600);color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer;">Importeer {{ aantal_ok }} bedrijven</button>
    <a href="/contacten/importeren" style="margin-left:10px;font-size:12.5px;color:var(--gray-400);text-decoration:none;">Annuleren</a>
</form>
{% else %}
<a href="/contacten/importeren" style="font-size:12.5px;color:var(--gray-400);text-decoration:none;">← Opnieuw proberen</a>
{% endif %}
{% endif %}
    """
    te_importeren = [r["data"] for r in preview_rijen if r["status"] == "ok"] if preview_rijen else []
    pagina = render_simple_page("Bedrijven importeren", "contacten", inhoud)
    return render_template_string(pagina, fout=fout, preview_rijen=preview_rijen, aantal_ok=aantal_ok, te_importeren=te_importeren)

@contacten_bp.route("/contacten/importeren/bevestigen", methods=["POST"])
def contacten_importeren_bevestigen():
    """Voert de daadwerkelijke import uit — pas nadat de preview is bevestigd."""
    _guard = vereist_afdeling_of_403("contacten")
    if _guard: return _guard

    import json as _json
    try:
        te_importeren = _json.loads(request.form.get("import_data", "[]"))
    except (ValueError, TypeError):
        te_importeren = []

    bestaande_namen_laag = {b["naam"].strip().lower() for b in ENF_BEDRIJVEN}
    status_alle = laad_status()
    accountmanagers_alle = laad_accountmanagers()
    personen = laad_contactpersonen()

    aantal_geimporteerd = 0
    for item in te_importeren:
        naam = (item.get("naam") or "").strip()
        if not naam or naam.lower() in bestaande_namen_laag:
            continue  # dubbele veiligheidscheck — nooit een bestaand bedrijf overschrijven
        bestaande_namen_laag.add(naam.lower())

        locatie = geocode_adres(item.get("land",""), item.get("regio",""))
        ENF_BEDRIJVEN.append({
            "naam": naam, "url": "", "regio": item.get("regio",""), "land": item.get("land",""),
            "klanttype": "", "materialen": item.get("materialen",""), "volume": "",
            "lat": locatie["lat"] if locatie else None, "lon": locatie["lon"] if locatie else None,
        })

        if item.get("status") in ("klant", "leverancier"):
            status_alle[naam] = item["status"]
        if item.get("accountmanager"):
            accountmanagers_alle[naam] = item["accountmanager"]
        if item.get("contact_naam"):
            personen.append({
                "id": str(uuid.uuid4()), "naam": item["contact_naam"], "bedrijf": naam,
                "rol": item.get("contact_functie",""), "email": item.get("contact_email",""),
                "telefoon": item.get("contact_telefoon",""), "laatst": datetime.date.today().isoformat(),
                "gebruiker": session.get("gebruikersnaam",""), "aangemaakt": datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
            })
        aantal_geimporteerd += 1

    bewaar_bedrijven()
    bewaar_status(status_alle)
    bewaar_accountmanagers(accountmanagers_alle)
    bewaar_contactpersonen(personen)

    inhoud = f"""
<div class="page-title">Import voltooid</div>
<div style="max-width:480px;background:#f0fdf4;border-radius:10px;padding:20px;">
    <div style="font-size:15px;font-weight:700;color:#16a34a;margin-bottom:6px;">✅ {aantal_geimporteerd} bedrijven geïmporteerd</div>
    <div style="font-size:12.5px;color:var(--gray-500);">Ze staan nu in het systeem en zijn direct doorzoekbaar.</div>
</div>
<a href="/contacten/importeren" style="display:inline-block;margin-top:16px;font-size:12.5px;color:var(--brand-600);text-decoration:none;">Nog een bestand importeren</a>
    """
    pagina = render_simple_page("Import voltooid", "contacten", inhoud)
    return render_template_string(pagina)

@contacten_bp.route("/contacten/nieuw")
def contacten_nieuw_keuze():
    """Startpunt van het aanmaken van een contactpersoon: eerst kiezen tussen een
    nieuw bedrijf (bedrijf en contact tegelijk aanmaken) of een bestaand bedrijf
    (extra contactpersoon toevoegen, of een bestaande vervangen)."""
    _guard = vereist_afdeling_of_403("contacten")
    if _guard: return _guard

    inhoud = """
<div style="font-size:12px;color:var(--gray-400);margin-bottom:6px;">
    <a href="/contacten" style="color:var(--gray-400);text-decoration:none;">Contacten</a> &nbsp;/&nbsp; <span style="color:var(--gray-600);">Nieuw</span>
</div>
<div class="page-title">Contactpersoon toevoegen</div>
<p style="color:var(--gray-400);margin-top:0;margin-bottom:24px;font-size:0.85rem;">Hoort deze persoon bij een bedrijf dat al in het systeem staat, of bij een nieuw bedrijf?</p>

<div style="display:flex;gap:16px;max-width:600px;">
    <a href="/contacten/nieuw/bedrijf" style="flex:1;display:block;padding:20px;border:1px solid var(--gray-200);border-radius:10px;text-decoration:none;color:inherit;">
        <div style="font-weight:700;color:var(--gray-800);font-size:14px;margin-bottom:6px;">Nieuw bedrijf</div>
        <div style="font-size:12.5px;color:var(--gray-500);">Het bedrijf staat nog niet in het systeem — bedrijf en contactpersoon in één keer aanmaken.</div>
    </a>
    <a href="/contacten/nieuw/bestaand" style="flex:1;display:block;padding:20px;border:1px solid var(--gray-200);border-radius:10px;text-decoration:none;color:inherit;">
        <div style="font-weight:700;color:var(--gray-800);font-size:14px;margin-bottom:6px;">Bestaand bedrijf</div>
        <div style="font-size:12.5px;color:var(--gray-500);">Een extra contactpersoon toevoegen bij een bedrijf dat al bestaat, of een bestaande vervangen.</div>
    </a>
</div>
<a href="/contacten/importeren" style="display:block;max-width:600px;margin-top:16px;padding:20px;border:1px solid var(--gray-200);border-radius:10px;text-decoration:none;color:inherit;">
    <div style="font-weight:700;color:var(--gray-800);font-size:14px;margin-bottom:6px;">📥 Bedrijven importeren via Excel</div>
    <div style="font-size:12.5px;color:var(--gray-500);">Meerdere bedrijven tegelijk toevoegen — handig bij een migratie vanuit een ander systeem (bv. Zoho).</div>
</a>
    """
    pagina = render_simple_page("Contact toevoegen", "contacten", inhoud)
    return render_template_string(pagina)

@contacten_bp.route("/contacten/nieuw/bedrijf", methods=["GET", "POST"])
def contacten_nieuw_bedrijf():
    """Nieuw bedrijf + contactpersoon in één keer aanmaken."""
    _guard = vereist_afdeling_of_403("contacten")
    if _guard: return _guard

    fout = None
    if request.method == "POST":
        bedrijfsnaam = request.form.get("bedrijfsnaam", "").strip()
        contactnaam = request.form.get("naam", "").strip()
        if not bedrijfsnaam or not contactnaam:
            fout = "Bedrijfsnaam en naam van de contactpersoon zijn verplicht."
        elif any(b["naam"].strip().lower() == bedrijfsnaam.lower() for b in ENF_BEDRIJVEN):
            fout = f"'{bedrijfsnaam}' bestaat al als bedrijf — kies bij 'Bestaand bedrijf' om daar een contactpersoon aan toe te voegen."
        else:
            nieuw_bedrijf = {
                "naam": bedrijfsnaam, "url": "", "regio": request.form.get("regio", "").strip(),
                "land": request.form.get("land", "").strip(), "klanttype": "",
                "materialen": request.form.get("materialen", "").strip(), "volume": "",
                "lat": None, "lon": None,
            }
            ENF_BEDRIJVEN.append(nieuw_bedrijf)
            bewaar_bedrijven()

            gekozen_status = request.form.get("status", "").strip()
            if gekozen_status in ("leverancier", "klant"):
                status_alle = laad_status()
                status_alle[bedrijfsnaam] = gekozen_status
                bewaar_status(status_alle)

            personen = laad_contactpersonen()
            personen.append({
                "id": str(uuid.uuid4()), "naam": contactnaam, "bedrijf": bedrijfsnaam,
                "rol": request.form.get("rol", "").strip(), "email": request.form.get("email", "").strip(),
                "telefoon": request.form.get("telefoon", "").strip(),
                "laatst": datetime.date.today().isoformat(),
                "gebruiker": session.get("gebruikersnaam", ""),
                "aangemaakt": datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
            })
            bewaar_contactpersonen(personen)
            return redirect(url_for("contacten.contacten"))

    inhoud = """
<div style="font-size:12px;color:var(--gray-400);margin-bottom:6px;">
    <a href="/contacten" style="color:var(--gray-400);text-decoration:none;">Contacten</a> &nbsp;/&nbsp;
    <a href="/contacten/nieuw" style="color:var(--gray-400);text-decoration:none;">Nieuw</a> &nbsp;/&nbsp; <span style="color:var(--gray-600);">Nieuw bedrijf</span>
</div>
<div class="page-title">Nieuw bedrijf + contactpersoon</div>
{% if fout %}<div style="background:#fef2f2;color:#dc2626;padding:10px 14px;border-radius:8px;margin-bottom:16px;font-size:12.5px;">{{ fout }}</div>{% endif %}

<form method="POST" style="max-width:520px;">
    <div style="font-size:11px;font-weight:700;color:var(--gray-400);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">Bedrijf</div>
    <input type="text" name="bedrijfsnaam" placeholder="Bedrijfsnaam" required style="width:100%;padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;box-sizing:border-box;margin-bottom:10px;">
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px;">
        <input type="text" name="land" placeholder="Land (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
        <input type="text" name="regio" placeholder="Regio/stad (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
    </div>
    <input type="text" name="materialen" placeholder="Materialen (optioneel)" style="width:100%;padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;box-sizing:border-box;margin-bottom:10px;">
    <select name="status" style="width:100%;padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;box-sizing:border-box;margin-bottom:20px;">
        <option value="leverancier">Leverancier</option>
        <option value="klant">Klant</option>
        <option value="">Geen van beide — alleen contactpersoon vastleggen</option>
    </select>

    <div style="font-size:11px;font-weight:700;color:var(--gray-400);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">Contactpersoon</div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px;">
        <input type="text" name="naam" placeholder="Naam" required style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
        <input type="text" name="rol" placeholder="Rol (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px;">
        <input type="email" name="email" placeholder="E-mail (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
        <input type="text" name="telefoon" placeholder="Telefoon (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
    </div>
    <button type="submit" style="padding:9px 18px;background:var(--brand-600);color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer;">Bedrijf en contact aanmaken</button>
</form>
    """
    pagina = render_simple_page("Nieuw bedrijf", "contacten", inhoud)
    return render_template_string(pagina, fout=fout)

@contacten_bp.route("/contacten/nieuw/bestaand", methods=["GET", "POST"])
def contacten_nieuw_bestaand():
    """Extra contactpersoon toevoegen bij een bestaand bedrijf, of een bestaande
    contactpersoon van dat bedrijf vervangen."""
    _guard = vereist_afdeling_of_403("contacten")
    if _guard: return _guard

    fout = None
    if request.method == "POST":
        bedrijfsnaam = request.form.get("bedrijfsnaam", "").strip()
        modus = request.form.get("modus", "extra")
        contactnaam = request.form.get("naam", "").strip()
        if not bedrijfsnaam or not contactnaam:
            fout = "Bedrijf en naam van de contactpersoon zijn verplicht."
        else:
            personen = laad_contactpersonen()
            nieuwe_gegevens = {
                "naam": contactnaam, "bedrijf": bedrijfsnaam,
                "rol": request.form.get("rol", "").strip(), "email": request.form.get("email", "").strip(),
                "telefoon": request.form.get("telefoon", "").strip(),
                "laatst": datetime.date.today().isoformat(),
                "gebruiker": session.get("gebruikersnaam", ""),
                "aangemaakt": datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
            }
            if modus == "vervangen":
                te_vervangen_id = request.form.get("te_vervangen_id", "")
                doel = next((p for p in personen if p["id"] == te_vervangen_id), None)
                if doel:
                    doel.update(nieuwe_gegevens)
                    bewaar_contactpersonen(personen)
                    return redirect(url_for("contacten.contacten"))
                fout = "De te vervangen contactpersoon is niet gevonden."
            else:
                nieuwe_gegevens["id"] = str(uuid.uuid4())
                personen.append(nieuwe_gegevens)
                bewaar_contactpersonen(personen)
                return redirect(url_for("contacten.contacten"))

    gekozen_bedrijf = request.args.get("bedrijf", "").strip()
    # Alleen bedrijven die al bij Leveranciers of Klanten staan — dus daadwerkelijk
    # een toegekende status óf accountmanager hebben, net als op die pagina's.
    # Geen doorzoeking van de volledige, ongefilterde bedrijvendatabase.
    status_alle = laad_status()
    accountmanagers_alle = laad_accountmanagers()
    bedrijfnamen = sorted({
        b["naam"] for b in list(ENF_BEDRIJVEN) + list(PAPIERFABRIEKEN)
        if status_alle.get(b["naam"]) or accountmanagers_alle.get(b["naam"])
    })
    personen_bij_bedrijf = [p for p in laad_contactpersonen() if p.get("bedrijf","").strip().lower() == gekozen_bedrijf.strip().lower()] if gekozen_bedrijf else []

    inhoud = """
<div style="font-size:12px;color:var(--gray-400);margin-bottom:6px;">
    <a href="/contacten" style="color:var(--gray-400);text-decoration:none;">Contacten</a> &nbsp;/&nbsp;
    <a href="/contacten/nieuw" style="color:var(--gray-400);text-decoration:none;">Nieuw</a> &nbsp;/&nbsp; <span style="color:var(--gray-600);">Bestaand bedrijf</span>
</div>
<div class="page-title">Contactpersoon bij een bestaand bedrijf</div>
{% if fout %}<div style="background:#fef2f2;color:#dc2626;padding:10px 14px;border-radius:8px;margin-bottom:16px;font-size:12.5px;">{{ fout }}</div>{% endif %}

<form method="GET" style="max-width:520px;margin-bottom:24px;">
    <div style="font-size:11px;font-weight:700;color:var(--gray-400);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">Bedrijf kiezen</div>
    <div style="display:flex;gap:8px;">
        <input type="text" name="bedrijf" value="{{ gekozen_bedrijf }}" list="bedrijvenLijstBestaand" placeholder="Begin te typen..." required style="flex:1;padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
        <datalist id="bedrijvenLijstBestaand">{% for naam in bedrijfnamen %}<option value="{{ naam }}">{% endfor %}</datalist>
        <button type="submit" style="padding:8px 16px;background:#fff;color:var(--gray-700);border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;">Verder</button>
    </div>
</form>

{% if gekozen_bedrijf %}
<form method="POST" style="max-width:520px;">
    <input type="hidden" name="bedrijfsnaam" value="{{ gekozen_bedrijf }}">
    <div style="font-size:11px;font-weight:700;color:var(--gray-400);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">{{ gekozen_bedrijf }}</div>

    <label style="display:flex;align-items:center;gap:8px;font-size:13px;color:var(--gray-700);margin-bottom:8px;cursor:pointer;">
        <input type="radio" name="modus" value="extra" checked onchange="document.getElementById('vervangKeuze').style.display='none';">
        Extra contactpersoon toevoegen
    </label>
    {% if personen_bij_bedrijf %}
    <label style="display:flex;align-items:center;gap:8px;font-size:13px;color:var(--gray-700);margin-bottom:12px;cursor:pointer;">
        <input type="radio" name="modus" value="vervangen" onchange="document.getElementById('vervangKeuze').style.display='block';">
        Een bestaande contactpersoon vervangen
    </label>
    <div id="vervangKeuze" style="display:none;margin-bottom:16px;padding:10px 12px;background:var(--gray-50);border-radius:8px;">
        {% for p in personen_bij_bedrijf %}
        <label style="display:flex;align-items:center;gap:8px;font-size:12.5px;color:var(--gray-600);padding:4px 0;cursor:pointer;">
            <input type="radio" name="te_vervangen_id" value="{{ p.id }}">
            {{ p.naam }}{% if p.rol %} — {{ p.rol }}{% endif %}
        </label>
        {% endfor %}
    </div>
    {% else %}
    <div style="font-size:12px;color:var(--gray-300);margin-bottom:16px;">Nog geen bestaande contactpersonen bij dit bedrijf om te vervangen.</div>
    {% endif %}

    <div style="font-size:11px;font-weight:700;color:var(--gray-400);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:10px;">Gegevens</div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px;">
        <input type="text" name="naam" placeholder="Naam" required style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
        <input type="text" name="rol" placeholder="Rol (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px;">
        <input type="email" name="email" placeholder="E-mail (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
        <input type="text" name="telefoon" placeholder="Telefoon (optioneel)" style="padding:8px 10px;border:1px solid var(--gray-200);border-radius:6px;font-size:13px;font-family:inherit;">
    </div>
    <button type="submit" style="padding:9px 18px;background:var(--brand-600);color:#fff;border:none;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer;">Opslaan</button>
</form>
{% endif %}
    """
    pagina = render_simple_page("Bestaand bedrijf", "contacten", inhoud)
    return render_template_string(pagina, fout=fout, gekozen_bedrijf=gekozen_bedrijf, bedrijfnamen=bedrijfnamen,
                                    personen_bij_bedrijf=personen_bij_bedrijf)